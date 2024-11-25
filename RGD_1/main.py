import pytesseract
from PIL import Image, ImageEnhance
import cv2
from cv2 import dnn_superres
from openpyxl import load_workbook
#
# # Создаём sr-объект
# sr = dnn_superres.DnnSuperResImpl_create()
#
# # Считываем изображение
# image = cv2.imread('test_3.png')
#
# # Считываем модель
# path = "FSRCNN_x2.pb"
# sr.readModel(path)
#
# # Устанавливаем модель и масштаб
# sr.setModel("fsrcnn", 3)
#
# # Улучшаем
# result = sr.upsample(image)
#
# # Сохраняем
# cv2.imwrite("test_3.png", result)
i = Image.open("test_3.png")
im = i.convert('L')
im = ImageEnhance.Brightness(im).enhance(1.15)
st = pytesseract.image_to_string(im, lang="rus")
print(st.split())
# print(st.split()[16],st.split()[17],st.split()[26],st.split()[28],st.split()[33],st.split()[37],st.split()[39],st.split()[43])
im = im.rotate(90, expand=True)
im = im.crop((600, 0, 2000, 300))
st += str(float(pytesseract.image_to_string(im).split()[-1]))

workbook = load_workbook("tabl.xlsx")
sheet = workbook.active
c = 0
k = 0
for i in range(2, 3):
    for j in range(1, 5):
        cell_value = sheet.cell(row=i, column=j).value
        if str(cell_value).upper() in st:
            c+=1
            k=i
if c==4 and sheet.cell(row=k, column=5).value != "Есть":
    print("YES")
    sheet.cell(row=k, column=5).value = "Есть"
else:
    print("LOX")

workbook.save("tabl.xlsx")
