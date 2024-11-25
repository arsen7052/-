from openpyxl import load_workbook

# Загрузка файла Excel
workbook = load_workbook("tabl.xlsx")

# Выбор листа для работы
sheet = workbook.active
for i in range(2,3):
    for j in range(1, 5):
        cell_value = sheet.cell(row=i, column=j).value
        print(str(cell_value))
